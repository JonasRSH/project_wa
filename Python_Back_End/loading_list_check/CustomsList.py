from pypdf import PdfReader
import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import date
from io import BytesIO


# Class which represents a shipment and handles the containing data
class Shipment():
    def __init__(self, position, shipment_no, exporter, colli_no, colli_type, content, weight, customs_handling):
        self.position = position
        self.shipment_no = shipment_no
        self.exporter = exporter
        self.colli_no = colli_no
        self.colli_type = colli_type
        self.content = content
        self.weight = weight
        self.customs_handling = customs_handling
            
            
    # Opens the PDF Shipment-List (Abmeldeliste) with the pypdf Library
    @staticmethod
    def open_checklist(pdf_path):
        try:
            reader = PdfReader(pdf_path, 'r')
            return reader, None
        except(FileNotFoundError):
            return None, 'File not Found'
        

    # Read and clean up the PDF data, which is a List of shipments with several pages and shipment informations
    @staticmethod
    def read_pdf_data(pdf_path):
        shipment_data_list = []
        reader, error_message = Shipment.open_checklist(pdf_path)
        if reader is None:
            return []
        for page in reader.pages:
            checklist_data = []
            data_from_pdf = page.extract_text()
            if not data_from_pdf:
                continue  # Leere Seite überspringen
            pdf_data = re.sub(r'^[-\s]+$', '', data_from_pdf, flags=re.MULTILINE)  # Entfernt die Trennlinien
            data_list = pdf_data.split('\n\n')

            for i in data_list:
                if i.startswith('0'):
                    checklist_data.append(i)
            for data in checklist_data:
                shipment_data = data.replace('     ',',').replace('\n', ',').replace('Bf','kg,').split('\n')
                for pdf_data in shipment_data:
                    shipment_data_list.append(pdf_data)
        return shipment_data_list

    #Extract list position from data
    @staticmethod
    def get_position(d):
        position = (d[0:3])
        return position

    #Extract shipment number from data
    @staticmethod
    def get_shipment_no(d):
        match = re.search(r'\d{11}', d)
        shipment_no = match.group(0) if match else None
        return shipment_no

    #Extract exporter from data
    @staticmethod
    def get_exporter(d):
        match = re.search(r'\d{11}\s+(\w+)', d)
        exporter = match.group(1) if match else None
        return exporter
        
    #Extract colli number from data
    @staticmethod
    def get_colli_no(d):
        colli = 'KT|EW|EU|KI|ZK|GB|PA|BX|DR|PH|HE'
        pattern = rf'(0[0-9]{{2}})(.*,)([0-9]{{1,3}}) ({colli})'
        matches = re.finditer(pattern, d)
        colli_no = 0
        for match in matches:
            collies = int(match.group(3))
            colli_no += collies
            return colli_no

    #Extract colli type from data
    @staticmethod
    def get_colli_type(d):
        colli = 'KT|EW|EU|KI|ZK|GB|PA|BX|DR|PH|HE'
        pattern = rf'(0[0-9]{{2}})(.*,)([0-9]{{1,3}}) ({colli})'
        match = re.search(pattern, d)
        if match:
            colli_type = match.group(4)
            return colli_type


    #Extract content from data
    @staticmethod
    def get_content(d):
        content_list = []
        content_data = d[45:100].split(' ')
        for content in content_data:
            if content.isalpha():
                if len(content) > 2:
                    content_list.append(content)
                    return content


    #Extract weight from data
    @staticmethod
    def get_weight(d):
        weight_pattern = rf'([0-9]{{1,5}}) (kg)'
        weight_match = re.finditer(weight_pattern, d)
        weight = 0
        for match in weight_match:
            weights = int(match.group(1))
            weight += weights
            return weight

    #Extracts customs handling from data
    @staticmethod
    def get_customs_handling(d):
        tokens = re.findall(r'EDEC|M90|EUR1|EU-VZ|EUVZ|GVZ|ATA|E-T1|S-T1', d)
        if tokens:
            return tokens
        return ['! No customs handling ! Please verify...']

    #Return True if the shipment is valid (not 'ZK'), False otherwise
    @staticmethod
    def is_valid_shipment(colli_type):
        return colli_type != 'ZK'
    
    #Calculates the sum of all Collies
    @staticmethod
    def calculate_total_collies(shipment_list):
        total_collies = sum(shipment.colli_no for shipment in shipment_list)
        return total_collies
    
    @staticmethod
    def calculate_total_weight(shipment_list):
        total_weight = sum(shipment.weight for shipment in shipment_list)
        return total_weight

    #Creates a shipment with the from the PDF extracted data
    @classmethod
    def create_shipment(cls, pdf_path):
        shipment_data = cls.read_pdf_data(pdf_path)
        shipment_list = []
        for d in shipment_data:
            shipment = cls(
                position=cls.get_position(d),
                shipment_no=cls.get_shipment_no(d),
                exporter=cls.get_exporter(d),
                colli_no=cls.get_colli_no(d),
                colli_type=cls.get_colli_type(d),
                content=cls.get_content(d),
                weight=cls.get_weight(d),
                customs_handling=cls.get_customs_handling(d)
            )
            if cls.is_valid_shipment(shipment.colli_type):
                shipment_list.append(shipment)
        return shipment_list
    
    def __str__(self):
        return (
            f'Shipment No: {self.shipment_no} | '
            f'Colli No: {self.colli_no} | '
            f'Colli Type: {self.colli_type} | '
            f'Content: {self.content} | '
            f'Weight: {self.weight} | '
            f'Customs Handling: {self.customs_handling}'
        )


    #Opens the Excel-file 'warenausweis.xlsx', writes the data and saves the file as a new copy in memory 
    def create_excel(self, filename=None, abfahrt_name=None, datum=None, kennzeichen=None, anhaenger=None, zollamt_abgang=None, zollamt_grenz=None):
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        # Vorlage immer gleich, nie überschreiben
        vorlage = os.path.join(base_dir, 'wa_automater', 'main', 'resources', 'warenausweis.xlsx')
        if abfahrt_name is None:
            abfahrt_name = ''
        if datum is None:
            from datetime import datetime
            datum = datetime.now().strftime('%Y-%m-%d')
        if kennzeichen is None:
            kennzeichen = ''
        if anhaenger is None:
            anhaenger = ''
        if zollamt_abgang is None:
            zollamt_abgang = ''
        if zollamt_grenz is None:
            zollamt_grenz = ''
        neuer_name = f"WA {abfahrt_name} {datum}.xlsx"
        # Entfernt problematische Zeichen aus dem Dateinamen
        neuer_name = re.sub(r'[\\/:"*?<>|()]+', '', neuer_name)
        
        wb = load_workbook(vorlage)
        sheetnames = ['Tabelle 1', 'Tabelle1', 'Tabelle 2', 'Tabelle2', 'Tabelle 3', 'Tabelle3']
        print('DEBUG: Vorhandene Sheets:', wb.sheetnames)
        for sheetname in sheetnames:
            if sheetname in wb.sheetnames:
                print(f'DEBUG: Schreibe Daten in {sheetname}')
                ws = wb[sheetname]
                # Abfahrt in E1
                if abfahrt_name:
                    ws['E1'] = abfahrt_name
                # Motorwagen-Kennzeichen in H3
                if kennzeichen:
                    ws['H3'] = kennzeichen
                # Anhänger-Kennzeichen in H4
                if anhaenger:
                    ws['H4'] = anhaenger
                # Grenzzollamt in B1
                if zollamt_grenz:
                    ws['B1'] = zollamt_grenz
                # Datum in B4
                if datum:
                    ws['B4'] = datum
                # Abgangszollamt (B2) wird erst später eingetragen, daher hier nicht automatisch setzen
                start_row = 6
                start_col = 2
                merges_to_remove = []
                for merged_range in ws.merged_cells.ranges:
                    min_col, min_row, max_col, max_row = merged_range.bounds
                    if min_row >= start_row and min_col >= start_col:
                        merges_to_remove.append(merged_range)
                for merged_range in merges_to_remove:
                    ws.unmerge_cells(str(merged_range))

                # Sortierung: OBEN = E-T1 durchgehend, UNTEN = S-T1 / Grenzverzollung
                def is_et1_durchgehend(s):
                    tokens = s.customs_handling
                    return ('E-T1' in tokens
                            and 'GVZ' not in tokens
                            and 'EU-VZ' not in tokens
                            and 'EUVZ' not in tokens
                            and 'S-T1' not in tokens)

                top_group    = [s for s in self.shipment_list if is_et1_durchgehend(s)]
                bottom_group = [s for s in self.shipment_list if not is_et1_durchgehend(s)]

                def write_shipment(row, shipment):
                    # Excel-Daten kommen aus dem T1-Transitdokument.
                    ws.cell(row=row, column=start_col,     value=getattr(shipment, 'mrn', None) or '')
                    ws.cell(row=row, column=start_col + 1, value=getattr(shipment, 't1_colli_no', None))
                    ws.cell(row=row, column=start_col + 2, value=getattr(shipment, 't1_colli_type', None) or '')
                    ws.cell(row=row, column=start_col + 3, value=getattr(shipment, 't1_goods_description', None) or '')
                    ws.cell(row=row, column=start_col + 4, value=getattr(shipment, 't1_weight', None))

                current_row = start_row

                # OBEN: E-T1 durchgehend
                for shipment in top_group:
                    write_shipment(current_row, shipment)
                    current_row += 1

                # Trennzeilen (nur wenn beide Gruppen belegt)
                blue_font = Font(color='0070C0', bold=True)
                anummer_row = None
                if top_group and bottom_group:
                    current_row += 2                                           # N+1, N+2 leer
                    abgang_cell = ws.cell(row=current_row, column=start_col,
                                         value=zollamt_abgang)                 # N+3
                    abgang_cell.font = blue_font
                    current_row += 1
                    anummer_row = current_row                                  # N+4 (wird nach unten befüllt)
                    current_row += 3                                           # N+5, N+6 leer → current = N+7

                # UNTEN: S-T1 / Grenzverzollung
                for shipment in bottom_group:
                    write_shipment(current_row, shipment)
                    current_row += 1

                # A-Nummer Text mit tatsächlichen Excel-Zeilennummern
                if anummer_row is not None and bottom_group:
                    bottom_start_row = anummer_row + 3 - 5  # Offset: Excel-Zeile relativ zur Vorlage
                    bottom_end_row   = bottom_start_row + len(bottom_group) - 1
                    anummer_cell = ws.cell(row=anummer_row, column=start_col,
                                          value=f'A-Nummer siehe Zeile {bottom_start_row} bis {bottom_end_row}')
                    anummer_cell.font = blue_font
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer, neuer_name 


def main():
    pdf_path = "/Users/Jonas_1/Documents/Jonas/Informatik/Projekt_WA/Python_Back_End/loading_list_check/abmeldeliste.pdf" 
    shipment_list = Shipment.create_shipment(pdf_path)
    for shipment in shipment_list:
        if shipment_list:
            shipment_list[0].shipment_list = shipment_list
            shipment_list[0].create_excel()
    print(f'Total Collies: {Shipment.calculate_total_collies(shipment_list)} Total Weight: {Shipment.calculate_total_weight(shipment_list)}')
    

if __name__ == "__main__":
    main()